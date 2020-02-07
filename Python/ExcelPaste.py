import openpyxl
import os

wb = openpyxl.load_workbook(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\20200205_強制地図位置解除対象リスト_61-80.xlsx')
#  wb = openpyxl.load_workbook(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\20200205_強制地図位置解除対象リスト_81-100.xlsx')
#  wb = openpyxl.load_workbook(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\20200205_強制地図位置解除対象リスト_100-108.xlsx')

#Debug用
print(wb.sheetnames)

#  設定値
FILE_ROOTPATH = r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_QA'

ws1  = wb['1']
ws2  = wb['2']
ws3  = wb['3']
ws4  = wb['4']
ws5  = wb['5']
ws6  = wb['6']
ws7  = wb['7']
ws8  = wb['8']
ws9  = wb['9']
ws10  = wb['10']
ws11  = wb['11']
ws12  = wb['12']
ws13  = wb['13']
ws14  = wb['14']
ws15  = wb['15']
ws16  = wb['16']
ws17  = wb['17']
ws18  = wb['18']
ws19  = wb['19']
ws20  = wb['20']


img1  = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '47605.PNG'))
img2  = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '47734.PNG'))
img3  = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '47758.PNG'))
img4  = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '48076.PNG'))
img5  = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '48178.PNG'))
img6  = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '48307.PNG'))
img7  = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '48478.PNG'))
img8  = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '48538.PNG'))
img9  = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '48568.PNG'))
img10 = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '48790.PNG'))
img11 = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '48823.PNG'))
img12 = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '48937.PNG'))
img13 = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '49024.PNG'))
img14 = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '49039.PNG'))
img15 = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '49036.PNG'))
img16 = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '49069.PNG'))
img17 = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '49279.PNG'))
img18 = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '68995.PNG'))
img19 = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '69178.PNG'))
img20 = openpyxl.drawing.image.Image(os.path.join(FILE_ROOTPATH, '69418.PNG'))

ws1.add_image( img1, 'A2' )
ws2.add_image( img2, 'A2' )
ws3.add_image( img3, 'A2' )
ws4.add_image( img4, 'A2' )
ws5.add_image( img5, 'A2' )
ws6.add_image( img6, 'A2' )
ws7.add_image( img7, 'A2' )
ws8.add_image( img8, 'A2' )
ws9.add_image( img9, 'A2' )
ws10.add_image( img10, 'A2' )
ws11.add_image( img11, 'A2' )
ws12.add_image( img12, 'A2' )
ws13.add_image( img13, 'A2' )
ws14.add_image( img14, 'A2' )
ws15.add_image( img15, 'A2' )
ws16.add_image( img16, 'A2' )
ws17.add_image( img17, 'A2' )
ws18.add_image( img18, 'A2' )
ws19.add_image( img19, 'A2' )
ws20.add_image( img20, 'A2' )

img1  = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\47605.PNG')
img2  = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\47734.PNG')
img3  = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\47758.PNG')
img4  = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\48076.PNG')
img5  = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\48178.PNG')
img6  = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\48307.PNG')
img7  = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\48478.PNG')
img8  = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\48538.PNG')
img9  = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\48568.PNG')
img10 = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\48790.PNG')
img11 = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\48823.PNG')
img12 = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\48937.PNG')
img13 = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\49024.PNG')
img14 = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\49039.PNG')
img15 = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\49036.PNG')
img16 = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\49069.PNG')
img17 = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\49279.PNG')
img18 = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\68995.PNG')
img19 = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\69178.PNG')
img20 = openpyxl.drawing.image.Image(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\Image_Hon\69418.PNG')

ws1.add_image( img1, 'A48' )
ws2.add_image( img2, 'A48' )
ws3.add_image( img3, 'A48' )
ws4.add_image( img4, 'A48' )
ws5.add_image( img5, 'A48' )
ws6.add_image( img6, 'A48' )
ws7.add_image( img7, 'A48' )
ws8.add_image( img8, 'A48' )
ws9.add_image( img9, 'A48' )
ws10.add_image( img10, 'A48' )
ws11.add_image( img11, 'A48' )
ws12.add_image( img12, 'A48' )
ws13.add_image( img13, 'A48' )
ws14.add_image( img14, 'A48' )
ws15.add_image( img15, 'A48' )
ws16.add_image( img16, 'A48' )
ws17.add_image( img17, 'A48' )
ws18.add_image( img18, 'A48' )
ws19.add_image( img19, 'A48' )
ws20.add_image( img20, 'A48' )


wb.save(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\20200205_強制地図位置解除対象リスト_61-80.xlsx')
#  wb.save(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\20200205_強制地図位置解除対象リスト_81-100.xlsx')
#  wb.save(r'C:\Users\setupuser\Documents\00_Ri\00_作業\70_定例作業\02_解除用\20200205_強制地図位置解除対象リスト_100-108.xlsx')
